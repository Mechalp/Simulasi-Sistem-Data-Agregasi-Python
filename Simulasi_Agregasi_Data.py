import os
import time
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Function to clear the screen
def clear_screen():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')

# Clear the screen
clear_screen()

# Wait for a moment before showing new content
time.sleep(1)

# Parameter Jaringan
xm = 100
ym = 100
n = 12
sinkx = 50
sinky = 50
Eo = 0.5
Eelec = 50 * 10**(-9)
ETx = 50 * 10**(-9)
ERx = 50 * 10**(-9)
Eamp = 100 * 10**(-12)
EDA = 5 * 10**(-9)
Esens = 5 * 10**(-9)
k = 200
p = 0.25
rnd = 0
operating_nodes = n
temp_val = 0
flag1stdead = 0
dead_nodes = 0
transmissions = 0

# Load or create WSN with fixed positions
if os.path.exists('wsn_cluster_info.xlsx'):
    df_positions = pd.read_excel('wsn_cluster_info.xlsx', sheet_name='Node Positions')
    SN = [{
        'id': int(row['Node ID']),
        'x': row['X'],
        'y': row['Y'],
        'E': Eo,
        'role': 0,
        'cluster': 0,
        'cond': 1,
        'rleft': 0,
        'dtch': 0,
        'dts': 0,
        'tel': 0,
        'rn': 0,
        'chid': 0,
        'active': True,
        'sensors': {
            'temperature': np.random.uniform(20, 35),
            'humidity': np.random.uniform(30, 90),
            'voltage': np.random.uniform(3.0, 4.2),
            'mq2': np.random.uniform(200, 10000)
        },
        'previous_sensors': {}
    } for index, row in df_positions.iterrows() if row['Node ID'] != 'Sink']
    sinkx = df_positions[df_positions['Node ID'] == 'Sink']['X'].values[0]
    sinky = df_positions[df_positions['Node ID'] == 'Sink']['Y'].values[0]
else:
    SN = [{
        'id': i,
        'x': np.random.rand() * xm,
        'y': np.random.rand() * ym,
        'E': Eo,
        'role': 0,
        'cluster': 0,
        'cond': 1,
        'rleft': 0,
        'dtch': 0,
        'dts': 0,
        'tel': 0,
        'rn': 0,
        'chid': 0,
        'active': True,
        'sensors': {
            'temperature': np.random.uniform(20, 35),
            'humidity': np.random.uniform(30, 90),
            'voltage': np.random.uniform(3.0, 4.2),
            'mq2': np.random.uniform(200, 10000)
        },
        'previous_sensors': {}
    } for i in range(1, n + 1)]
    positions = [{'Node ID': node['id'], 'X': node['x'], 'Y': node['y']} for node in SN]
    positions.append({'Node ID': 'Sink', 'X': sinkx, 'Y': sinky})

    # Save initial positions
    df_positions = pd.DataFrame(positions)
    with pd.ExcelWriter('wsn_cluster_info.xlsx') as writer:
        df_positions.to_excel(writer, sheet_name='Node Positions', index=False)

# Menetapkan cluster awal untuk setiap node
clusters = [1, 2, 3]
for i, node in enumerate(SN):
    node['cluster'] = clusters[i % len(clusters)]

clusterInfoPerRound = []
nrg = []
avg_node = []
total_energy_per_round = []
dead_nodes_per_round = []
rounds_first_dead = []
aggregated_data_per_round = []
non_aggregated_data_per_round = []
energy_consumption_per_node = []

# Fungsi untuk update sensor readings setiap ronde
def update_sensor_readings(node):
    node['sensors']['temperature'] = np.random.uniform(20, 35)
    node['sensors']['humidity'] = np.random.uniform(30, 90)
    node['sensors']['voltage'] = np.random.uniform(3.0, 4.2)
    node['sensors']['mq2'] = np.random.uniform(200, 10000)

def calculate_similarity(current_sensors, previous_sensors):
    if not previous_sensors:
        return False
    similarities = []
    for sensor in current_sensors:
        current = current_sensors[sensor]
        previous = previous_sensors.get(sensor, current)
        similarity = 1 - abs(current - previous) / max(abs(current), abs(previous), 1)
        similarities.append(similarity > 0.4)  # Adjusted similarity threshold to 0.4
    return all(similarities)

while operating_nodes > 0:
    rnd += 1
    t = p / (1 - p * (rnd % (1 / p)))
    tleft = rnd % (1 / p)
    CLheads = 0
    energy = 0
    clusterInfoThisRound = []
    CL = []  # Pastikan CL diinisialisasi ulang setiap ronde
    energy_consumed_this_round = []

    for node in SN:
        update_sensor_readings(node)
        node['role'] = 0
        node['chid'] = 0
        if node['rleft'] > 0:
            node['rleft'] -= 1

    # Memilih cluster head baru untuk setiap cluster
    unique_clusters = set(node['cluster'] for node in SN)
    for cluster in unique_clusters:
        cluster_members = [node for node in SN if node['cluster'] == cluster]
        new_ch = np.random.choice(cluster_members)
        new_ch['role'] = 1
        new_ch['chx'] = new_ch['x']
        new_ch['chy'] = new_ch['y']
        new_ch['rn'] = rnd
        new_ch['tel'] += 1
        new_ch['rleft'] = 1 / p - tleft
        new_ch['dts'] = np.sqrt((sinkx - new_ch['x'])**2 + (sinky - new_ch['y'])**2)
        new_ch['chid'] = new_ch['id']
        CLheads += 1

        cluster_info = {
            'Round': rnd,
            'Cluster': cluster,
            'CH': new_ch['id'],
            'Member': len(cluster_members) - 1,
            'Node Active': operating_nodes,
            'CH_ID': new_ch['id'],
            'Members_ID': [node['id'] for node in cluster_members if node['id'] != new_ch['id']],
            'x': new_ch['x'],
            'y': new_ch['y']
        }
        CL.append(cluster_info)
        clusterInfoThisRound.append(cluster_info)

    clusterInfoPerRound.append(clusterInfoThisRound)

    non_aggregated_data = []
    aggregated_data = []

    for node in SN:
        if node['cond'] == 1 and node['role'] == 0:
            current_sensors = node['sensors']
            similarity = calculate_similarity(current_sensors, node['previous_sensors'])
            if not similarity:
                data_size = sum([len(str(v)) for v in current_sensors.values()])
                ETx = Eelec * data_size * k + Eamp * data_size * k * node['dtch']**2
                ERx = Eelec * data_size * k
                Econs = ETx + ERx

                if node['E'] > Econs:
                    node['E'] -= Econs
                    energy += Econs
                    energy_consumed_this_round.append({
                        'Round': rnd,
                        'Node ID': node['id'],
                        'Role': 'Member',
                        'Energy Consumed': Econs
                    })
                else:
                    dead_nodes += 1
                    operating_nodes -= 1
                    node['active'] = False
                    node['cond'] = 0
                    node['chid'] = 0
                    node['rop'] = rnd

                node['previous_sensors'] = current_sensors.copy()

            row = [
                node['cluster'], node['id'],
                node['sensors']['temperature'], node['previous_sensors'].get('temperature', 0),
                node['sensors']['humidity'], node['previous_sensors'].get('humidity', 0),
                node['sensors']['voltage'], node['previous_sensors'].get('voltage', 0),
                node['sensors']['mq2'], node['previous_sensors'].get('mq2', 0),
                similarity
            ]
            if similarity:
                non_aggregated_data.append(row)
            else:
                aggregated_data.append(row)

    non_aggregated_data_per_round.append(non_aggregated_data)
    aggregated_data_per_round.append(aggregated_data)

    for i in range(n):
        if SN[i]['cond'] == 1 and SN[i]['role'] == 1:
            ETx = (Eelec + EDA) * k + Eamp * k * SN[i]['dts']**2
            ERx_total = (Eelec + EDA) * k * len([node for node in SN if node['cluster'] == SN[i]['cluster'] and node['role'] == 0])
            Econs = ETx + ERx_total
            if SN[i]['E'] > Econs:
                SN[i]['E'] -= Econs
                energy += Econs
                energy_consumed_this_round.append({
                    'Round': rnd,
                    'Node ID': SN[i]['id'],
                    'Role': 'CH',
                    'Energy Consumed': Econs
                })
            else:
                dead_nodes += 1
                operating_nodes -= 1
                SN[i]['active'] = False
                SN[i]['cond'] = 0
                SN[i]['rop'] = rnd

    energy_consumption_per_node.append(energy_consumed_this_round)

    if operating_nodes < n and temp_val == 0:
        temp_val = 1
        flag1stdead = rnd
    avg_energy_per_node = energy / operating_nodes if operating_nodes > 0 else 0
    avg_node.append(avg_energy_per_node)
    total_energy_consumed_this_round = sum(Eo - node['E'] for node in SN if node['cond'] == 1)
    total_energy_per_round.append(total_energy_consumed_this_round)
    dead_nodes_per_round.append(dead_nodes)
    if energy > 0:
        nrg.append(energy)
        transmissions += 1

    rounds_first_dead.append(flag1stdead)

# Save cluster information to first Excel file
cluster_data = []
for round_data in clusterInfoPerRound:
    for cluster_info in round_data:
        cluster_data.append(cluster_info)

df_cluster = pd.DataFrame(cluster_data)

with pd.ExcelWriter('wsn_cluster_info.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_cluster.to_excel(writer, sheet_name='Cluster Info', index=False)

    # Save energy consumption per node to new sheet
    energy_consumption_flat = [item for sublist in energy_consumption_per_node for item in sublist]
    df_energy_consumption = pd.DataFrame(energy_consumption_flat)
    df_energy_consumption.to_excel(writer, sheet_name='Energy Consumption', index=False)

# Save aggregation information to second Excel file
aggregation_info = {
    'Node pertama mati pada ronde': [flag1stdead],
    'Length of clusterInfoPerRound': [len(clusterInfoPerRound)]
}

df_aggregation_info = pd.DataFrame(aggregation_info)
df_aggregated = pd.DataFrame([item for sublist in aggregated_data_per_round for item in sublist], columns=[
    'Cluster ID', 'Node ID', 'Temp (Now)', 'Temp (Prev)', 'Hum (Now)', 'Hum (Prev)', 'Voltage (Now)', 'Voltage (Prev)', 'MQ2 (Now)', 'MQ2 (Prev)', 'Similarity (%)'])
df_non_aggregated = pd.DataFrame([item for sublist in non_aggregated_data_per_round for item in sublist], columns=[
    'Cluster ID', 'Node ID', 'Temp (Now)', 'Temp (Prev)', 'Hum (Now)', 'Hum (Prev)', 'Voltage (Now)', 'Voltage (Prev)', 'MQ2 (Now)', 'MQ2 (Prev)', 'Similarity (%)'])

# Mengubah nilai kolom "Similarity (%)" menjadi "Yes" atau "No"
df_aggregated['Similarity (%)'] = df_aggregated['Similarity (%)'].apply(lambda x: 'no' if x else 'yes')
df_non_aggregated['Similarity (%)'] = df_non_aggregated['Similarity (%)'].apply(lambda x: 'no' if x else 'yes')

with pd.ExcelWriter('wsn_aggregation_info.xlsx') as writer:
    df_aggregation_info.to_excel(writer, sheet_name='Info', index=False)

    # Split aggregated data into multiple sheets if necessary
    max_rows_per_sheet = 1000000
    for i in range(0, len(df_aggregated), max_rows_per_sheet):
        df_non_aggregated.iloc[i:i+max_rows_per_sheet].to_excel(writer, sheet_name=f'Non Aggregated Data {i // max_rows_per_sheet + 1}', index=False)

    # Split non-aggregated data into multiple sheets if necessary
    for i in range(0, len(df_non_aggregated ), max_rows_per_sheet):
       df_aggregated .iloc[i:i+max_rows_per_sheet].to_excel(writer, sheet_name=f'Aggregated Data {i // max_rows_per_sheet + 1}', index=False)

print('Simulasi selesai dan data telah disimpan dalam file wsn_cluster_info.xlsx dan wsn_aggregation_info.xlsx')

# Di luar loop simulasi
if flag1stdead > 0:
    print(f'Node pertama mati pada ronde {flag1stdead}.')
else:
    print('Tidak ada node yang mati selama simulasi.')
print("Length of clusterInfoPerRound:", len(clusterInfoPerRound))

round_to_check = int(input("Masukkan ronde yang ingin dicek: "))

if round_to_check > 0 and round_to_check <= rnd:
    non_aggregated_data = non_aggregated_data_per_round[round_to_check - 1]
    aggregated_data = aggregated_data_per_round[round_to_check - 1]

    print("Data yang diAgregasi:")
    print(f"{'Cluster ID':>10} {'Node ID':>10} {'Temp (Now)':>12} {'Temp (Prev)':>12} {'Hum (Now)':>10} {'Hum (Prev)':>10} {'Voltage (Now)':>10} {'Voltage (Prev)':>10} {'MQ2 (Now)':>10} {'MQ2 (Prev)':>10} {'Similarity (%)':>15} {'Aggregated':>10}")
    for row in aggregated_data:
        print(f"{row[0]:>10} {row[1]:>10} {row[2]:>12.2f} {row[3]:>12.2f} {row[4]:>10.2f} {row[5]:>10.2f} {row[6]:>10.2f} {row[7]:>10.2f} {row[8]:>10.2f} {row[9]:>10.2f} {'Yes':>10}")

    print("\nData yang tidak diAgregasi:")
    print(f"{'Cluster ID':>10} {'Node ID':>10} {'Temp (Now)':>12} {'Temp (Prev)':>12} {'Hum (Now)':>10} {'Hum (Prev)':>10} {'Voltage (Now)':>10} {'Voltage (Prev)':>10} {'MQ2 (Now)':>10} {'MQ2 (Prev)':>10} {'Similarity (%)':>15} {'Aggregated':>10}")
    for row in non_aggregated_data:
        print(f"{row[0]:>10} {row[1]:>10} {row[2]:>12.2f} {row[3]:>12.2f} {row[4]:>10.2f} {row[5]:>10.2f} {row[6]:>10.2f} {row[7]:>10.2f} {row[8]:>10.2f} {row[9]:>10.2f} {'No':>10}")
else:
    print("Ronde yang dimasukkan tidak valid atau belum tercapai.")

# Menyimpan dan memformat sheet "Energy Consumption" sesuai contoh
energy_consumption_data = []

for round_data in energy_consumption_per_node:
    if len(round_data) > 0:
        round_number = round_data[0]['Round']
        row = [round_number]
        for node_data in round_data:
            row.append(node_data['Node ID'])
            row.append(node_data['Role'])
            row.append(node_data['Energy Consumed'])
        energy_consumption_data.append(row)

# Ensure each Node column contains consistent data
nodes = list(set(node_data['Node ID'] for round_data in energy_consumption_per_node for node_data in round_data))

formatted_data = []
for i in range(rnd):
    round_number = i + 1
    row = [round_number]
    for node in nodes:
        node_info = next((data for data in energy_consumption_per_node[i] if data['Node ID'] == node), None)
        if node_info:
            row.extend([node_info['Node ID'], node_info['Role'], node_info['Energy Consumed']])
        else:
            row.extend([node, 'Member', 0])  # Default to 'Member' and 0 energy if no data
    formatted_data.append(row)

df_energy_consumption_formatted = pd.DataFrame(formatted_data)
df_energy_consumption_formatted.columns = ['Round'] + [f'{node}_{attr}' for node in nodes for attr in ['Node', 'Role', 'Energy Consumed']]

with pd.ExcelWriter('wsn_cluster_info.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_energy_consumption_formatted.to_excel(writer, sheet_name='Energy Consumption', index=False)

# Membuka kembali file dan menerapkan format
wb = load_workbook('wsn_cluster_info.xlsx')
ws = wb['Energy Consumption']

# Mengatur lebar kolom dan format angka
for column in ws.iter_cols(min_col=2, max_col=ws.max_column):
    for cell in column:
        if 'Energy Consumed' in cell.column_letter:
            cell.number_format = '0.00000'
            cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[column[0].column_letter].width = 15

wb.save('wsn_cluster_info.xlsx')

print("File Excel dengan format yang diinginkan berhasil dibuat.")

# Plotting Simulation Results
# Plot clustering information
plt.figure(figsize=(10, 10))

# Define cluster colors
cluster_colors = {1: 'blue', 2: 'orange', 3: 'green'}

# Plot member nodes with different colors for each cluster
for cluster_info in clusterInfoPerRound[round_to_check - 1]:
    cluster = cluster_info['Cluster']
    cluster_nodes = [node for node in SN if node['cluster'] == cluster and node['id'] != cluster_info['CH_ID']]
    plt.scatter([node['x'] for node in cluster_nodes], [node['y'] for node in cluster_nodes], c=cluster_colors[cluster], label=f'Cluster {cluster} Nodes', marker='o')

# Plot cluster heads with different colors for each cluster
for cluster_info in clusterInfoPerRound[round_to_check - 1]:
    ch_id = cluster_info['CH_ID']
    ch = next(node for node in SN if node['id'] == ch_id)
    plt.scatter(ch['x'], ch['y'], c=cluster_colors[cluster_info['Cluster']], s=100, label=f'Cluster {cluster_info["Cluster"]} Heads', marker='s')

# Plot the sink as a triangle
plt.scatter(sinkx, sinky, c='red', marker='^', s=250, label='Sink')

# Plot lines between nodes and their cluster head
for cluster_info in clusterInfoPerRound[round_to_check - 1]:
    ch_id = cluster_info['CH_ID']
    ch = next(node for node in SN if node['id'] == ch_id)
    for member_id in cluster_info['Members_ID']:
        member = next(node for node in SN if node['id'] == member_id)
        plt.plot([ch['x'], member['x']], [ch['y'], member['y']], c=cluster_colors[cluster_info['Cluster']], linestyle='--', alpha=0.4)

# Plot lines between cluster heads and the sink
for cluster_info in clusterInfoPerRound[round_to_check - 1]:
    ch_id = cluster_info['CH_ID']
    ch = next(node for node in SN if node['id'] == ch_id)
    plt.plot([ch['x'], sinkx], [ch['y'], sinky], c='black', linestyle='--')

plt.title(f'Cluster Formation in WSN (Round {round_to_check})')
plt.xlabel('X coordinate')
plt.ylabel('Y coordinate')
plt.legend()
plt.grid(True)
plt.show()

# Plot Energy consumed per Transmission
if len(nrg) > 0:  # Pastikan nrg tidak kosong
    plt.figure(figsize=(8, 6))
    plt.plot(range(1, transmissions + 1), nrg, '-b', linewidth=1)

    plt.xlabel('Transmission')
    plt.ylabel('Energy (Joule)')
    plt.title('Energy Consumed per Transmission')
    plt.grid(True)
    plt.show()

# Plot Average Energy consumed by a Node per Transmission
plt.figure(figsize=(8, 6))
plt.plot(range(1, transmissions + 1), avg_node[:transmissions], '-r', linewidth=1)
plt.xlabel('Transmission')
plt.ylabel('Energy (Joule)')
plt.title('Average Energy Consumed by a Node per Transmission')
plt.grid(True)
plt.show()

# Plot Total Energy Consumed by the WSN per Transmission
plt.figure(figsize=(8, 6))
plt.plot(range(1, transmissions + 1), total_energy_per_round[:transmissions], '-g', linewidth=1)
plt.xlabel('Transmission')
plt.ylabel('Energy (Joule)')
plt.title('Total Energy Consumed by the WSN per Transmission')
plt.grid(True)
plt.show()

# Plot Jumlah node mati per round
plt.figure(figsize=(8, 6))
plt.plot(range(1, rnd + 1), dead_nodes_per_round, '-m', linewidth=1)
plt.xlabel('Round')
plt.ylabel('Number of Dead Nodes')
plt.title('Number of Dead Nodes per Round')
plt.grid(True)
plt.show()

# Plot Jumlah Node yang aktiv atau beroperasi tiap round
plt.figure(figsize=(8, 6))
plt.plot(range(1, rnd + 1), [n - dead_nodes for dead_nodes in dead_nodes_per_round], '-c', linewidth=1)
plt.xlabel('Round')
plt.ylabel('Number of Active Nodes')
plt.title('Number of Active Nodes per Round')
plt.grid(True)
plt.show()
